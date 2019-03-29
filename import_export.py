from openpyxl import Workbook, load_workbook
from db_man import DatabaseManager
import wine_bottle
import os


def export_db(path):
    # Exports the database to an excel file at the specified
    # file path
    # First, create a workbook. Condensed uses Qty instead of location
    # Expanded preserves location data
    exp_wb = Workbook()
    exp_condensed = exp_wb.active
    exp_condensed.title = 'Condensed Inventory'
    exp_full = exp_wb.create_sheet('Expanded Inventory')

    # Setup the pages to have the right headers. Grab lists of the col names and merge them
    db_export = DatabaseManager()
    col_keys = db_export.db_getcolnames('winedata')
    col_keys.extend(db_export.db_getcolnames('userinventory')[1:])
    exp_full.append(col_keys)

    # Replace the 'location' tag with 'qty' and add it to the condensed table and remove the dates
    qty_index = col_keys.index('location')
    col_keys[qty_index] = 'qty'
    date_index = col_keys.index('date_in')
    del col_keys[date_index:date_index+2]
    exp_condensed.append(col_keys)

    # now, grab everything from each table and store it in a list of lists
    inv_rows = db_export.db_fetch('SELECT * FROM userinventory WHERE date_out IS NULL ORDER BY wine_id', rows='all')


    # Add everything to the excel file sequentially
    # First, initialize the variables that may not exist 
    # until later in the loop
    old_id = None
    qty = 0
    old_row = []

    # Iterate through each of the entries in the inventory For each
    # entry, look it up in the winedata table and add create one 
    # long list which the relevant data. This is a little slower,
    # but more memory efficient. Write that to the extended page. 
    for i in range(len(inv_rows)):
        write_row = list(db_export.db_fetch('SELECT * FROM winedata WHERE wine_id=?', (inv_rows[i][0],)))
        write_row.extend(inv_rows[i][1:])
        exp_full.append(write_row)
        # Remove the date in/out entries
        del write_row[date_index:date_index+2]
        # The condensed row will run one iteration behind in order
        # to check for another of the same bottle. It is initialized
        # as none to avoid a definition error in the first loop
        if old_row != []:
            # if the wine id is the same, increment the counter
            if old_id == inv_rows[i][0]:
                qty += 1
            # if it's not, incriment the counter to account for the
            # last bottle, write it, and reset qty counter
            else:
                qty += 1
                old_row[qty_index] = qty
                exp_condensed.append(old_row)
                qty = 0
        # set old row and old id so we can check it against the 
        # next index on the next cycle
        old_row = write_row.copy()
        old_id = inv_rows[i][0]
    # Because the condensed table runs one cycle behind, we have to
    # run the write command one more time to finish the last entry
    qty += 1
    old_row[qty_index] = qty
    exp_condensed.append(old_row)


    # Save the file
    exp_wb.save(path)

def import_db(path):
    # This function imports an excel file (from a specified format),
    # checks for duplicates, and enters the non-duplicates into the
    # database. (Note: duplicates refer to the winedata table, not
    # the inventory. That will allow any)
    
    # First, grab the col names
    db_import = DatabaseManager()
    wine_keys = db_import.db_getcolnames('winedata')
    bottle_keys = db_import.db_getcolnames('userinventory')

    # Create an instance of the bottle object
    bottle = wine_bottle.Bottle({}, {})

    import_wb = load_workbook(path)
    import_ws = import_wb.active

    # Grab all rows from the spreadsheet and make a dictionary with
    # them. This will be used to add the entries to the db
    import_rows = tuple(import_ws.rows)
    input_dict = {}
    for key in import_rows[0]:
        input_dict[key.value] = None

    for i in range(len(import_rows)-1):
        for j, key in enumerate(input_dict):
            input_dict[key] = import_rows[i+1][j].value

        for wine_key in wine_keys:
            bottle.wine_info[wine_key] = input_dict[wine_key] if wine_key in input_dict else None
            
        for bottle_key in bottle_keys:
            bottle.bottle_info[bottle_key] = input_dict[bottle_key] if bottle_key in input_dict else None
        
        if 'expanded' in path[-23:-5]:
            bottle.add_new()
            bottle.clear_bottle()
        elif 'condensed' in path[-23:-5]:
            for k in range(input_dict['qty']):
                bottle.add_new()
            bottle.clear_bottle()

def generate_sheet(path, expanded=False):
    # Generates a template to be filled and imported. Can accept either condensed or expanded
    temp_wb = Workbook()
    temp_sheet = temp_wb.active
    # Setup the pages to have the right headers. Grab lists of the col names and merge them
    db_export = DatabaseManager()
    col_keys = db_export.db_getcolnames('winedata')
    col_keys.extend(db_export.db_getcolnames('userinventory')[1:])

    if expanded:
        temp_sheet.title = 'Expanded Inventory'
        temp_sheet.append(col_keys)

        # Add a label to the filename so the parser knows which one we're looking at
        path += 'template_expanded.xlsx'

    else:
        temp_sheet.title = 'Condensed Inventory'

        # Replace the 'location' tag with 'qty' and add it to the condensed table and remove the dates
        qty_index = col_keys.index('location')
        col_keys[qty_index] = 'qty'
        date_index = col_keys.index('date_in')
        del col_keys[date_index:date_index+2]
        temp_sheet.append(col_keys)

        # Add a label to the filename so the parser knows which one we're looking at
        path += 'template_condensed.xlsx'
    
    # Save the file
    temp_wb.save(path)

# Test Code
if __name__ == "__main__":
    path = 'template_condensed.xlsx'
    # export_db(path)
    import_db(path)
    # generate_sheet(path, expanded=False)

